"""pg_tabledef.writer.excel: list[TableDef] → output/테이블정의서.xlsx.

레이아웃 구조 (테이블 1개 기준, template/sample.xlsx 기준):
  Row +1 : Table 명 / TableSpace / Sub System / 주제영역명 / 주제영역명약어
  Row +2 : Entity 명 / 최초작성일 / 최종수정일 / 엔티티분류 / 오너쉽
  Row +3 : Entity 정의 (높이 120, B:L 병합)
  Row +4 : Key List 헤더 (높이 40)
  Row +5 : Primary Key
  Row +6 : Foreign Key
  Row +7 : Index Key
  Row +8 : Sequence Key (빈칸)
  Row +9 : Partitions Key 헤더 (빈칸)
  Row +10: Partitions 값 (빈칸)
  Row +11: SubPartitions Key (빈칸)
  Row +12: 컬럼 헤더 Row 1 (병합 있음)
  Row +13: 컬럼 헤더 Row 2 (Attribute Type / Relation & Value 서브헤더)
  Row +14~: 컬럼 데이터
"""
from __future__ import annotations

import csv
import json
import os
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ..models import TableDef
from . import styles as S

_RULES_DIR = Path(__file__).parent.parent.parent / "rules"


def _load_column_rules() -> dict[tuple[str, str], tuple[str, str]]:
    """rules/column_attribute_rules.json 로드 → {(column_name, attribute_name): (attr_type, rel_val)}"""
    path = _RULES_DIR / "column_attribute_rules.json"
    if not path.exists():
        return {}
    with path.open(encoding="utf-8") as f:
        rules = json.load(f)
    return {
        (r["column_name"], r["attribute_name"]): (r["attr_type"], r["rel_val"])
        for r in rules
    }


def _load_dtl_code_rules() -> tuple[dict[str, list[str]], dict[str, tuple[str, list[str]]]]:
    """rules/dtl_code.csv 로드.

    Returns:
        by_name: {code_group: [code_value, ...]}          # B열(col.name) 매칭용
        by_attr: {description: (code_group, [code_value, ...])}  # C열(col.attribute_name) 매칭용
    """
    path = _RULES_DIR / "dtl_code.csv"
    if not path.exists():
        return {}, {}
    by_name: dict[str, list[str]] = {}
    by_attr: dict[str, tuple[str, list[str]]] = {}
    with path.open(encoding="utf-8") as f:
        for row in csv.reader(f):
            if len(row) >= 4:
                group, desc, code_val = row[0], row[1], row[3]
                by_name.setdefault(group, []).append(code_val)
                if desc not in by_attr:
                    by_attr[desc] = (group, [])
                by_attr[desc][1].append(code_val)
    return by_name, by_attr


def _load_subject_rules() -> dict:
    """rules/table_subject_rules.json 로드."""
    path = _RULES_DIR / "table_subject_rules.json"
    if not path.exists():
        return {"strip_prefixes": [], "rules": []}
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def _resolve_subject(table_name: str) -> tuple[str, str, str]:
    """테이블명 → (sub_system, subject_area, subject_area_abbr).

    strip_prefixes 제거 후 rules 배열 순서대로 첫 번째 매칭 반환.
    매칭 없으면 ("", "", "").
    """
    name = table_name.upper()
    for prefix in _SUBJECT_RULES.get("strip_prefixes", []):
        if name.startswith(prefix.upper()):
            name = name[len(prefix):]
            break
    for rule in _SUBJECT_RULES.get("rules", []):
        pattern = rule["pattern"].upper()
        matched = (
            name.startswith(pattern) if rule.get("type") == "prefix"
            else pattern in name
        )
        if matched:
            return (
                rule.get("sub_system", ""),
                rule.get("subject_area", ""),
                rule.get("subject_area_abbr", ""),
            )
    return ("", "", "")


_COLUMN_RULES = _load_column_rules()
_DTL_BY_NAME, _DTL_BY_ATTR = _load_dtl_code_rules()
_SUBJECT_RULES = _load_subject_rules()

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


class ExcelWriter:
    """list[TableDef]를 받아 엑셀 파일로 저장."""

    DEFAULT_OUTPUT_PATH = Path("output") / "테이블정의서.xlsx"

    def __init__(self, output_path: str | Path | None = None) -> None:
        self.output_path = Path(output_path) if output_path else self.DEFAULT_OUTPUT_PATH

    def write(self, tables: list[TableDef]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "테이블정의서"

        row = 1
        for table in tables:
            row = self._write_table_header(ws, row, table)
            row = self._write_key_list(ws, row, table)
            row = self._write_columns(ws, row, table)
            row += 2  # 테이블 간 빈 행 2개

        self._apply_column_widths(ws)

        os.makedirs(self.output_path.parent, exist_ok=True)
        wb.save(str(self.output_path))

    # ──────────────────────────────────────────────────────────────────────────
    # 섹션 1: 테이블 정보 헤더 (3행)
    # ──────────────────────────────────────────────────────────────────────────

    def _write_table_header(self, ws: "Worksheet", row: int, table: TableDef) -> int:
        # Row 1: Table 명 | {name} | TableSpace | [D:E blank] | Sub System | [H blank] | 주제영역명 | [J blank] | 주제영역명약어 | [L blank]
        r = row
        sub_system, subject_area, subject_area_abbr = _resolve_subject(table.name)
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        self._wc(ws, r, 1,  "Table 명",       align=S.ALIGN_CENTER,      fill=S.FILL_HEADER)
        self._wc(ws, r, 2,  table.name,        align=S.ALIGN_LEFT_NO_WRAP)
        self._wc(ws, r, 3,  "TableSpace",      align=S.ALIGN_CENTER,      fill=S.FILL_HEADER)
        self._wc(ws, r, 4,  "",                align=S.ALIGN_CENTER)  # TableSpace 값
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        self._wc(ws, r, 6,  "Sub System",      align=S.ALIGN_CENTER,      fill=S.FILL_HEADER)  # F:G 병합
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        self._wc(ws, r, 8,  sub_system,         align=S.ALIGN_CENTER)  # Sub System 값
        self._wc(ws, r, 9,  "주제영역명",       align=S.ALIGN_CENTER,      fill=S.FILL_HEADER)
        self._wc(ws, r, 10, subject_area,        align=S.ALIGN_CENTER)  # 주제영역명 값
        self._wc(ws, r, 11, "주제영역명약어",   align=S.ALIGN_CENTER,      fill=S.FILL_HEADER)
        self._wc(ws, r, 12, subject_area_abbr,   align=S.ALIGN_CENTER)  # 주제영역명약어 값
        r += 1

        # Row 2: Entity 명 | {comment} | 최초작성일 | [D:E blank] | 최종수정일 | [H blank] | 엔티티분류 | [J blank] | 오너쉽 | [L blank]
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        comment_font = S.FONT_AI_SUGGESTED if table.comment_ai else None
        self._wc(ws, r, 1,  "Entity 명",   align=S.ALIGN_CENTER,      fill=S.FILL_HEADER)
        self._wc(ws, r, 2,  table.comment, align=S.ALIGN_LEFT_NO_WRAP, font=comment_font)
        self._wc(ws, r, 3,  "최초작성일",   align=S.ALIGN_CENTER,      fill=S.FILL_HEADER)
        self._wc(ws, r, 4,  "",             align=S.ALIGN_CENTER)  # 최초작성일 값
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        self._wc(ws, r, 6,  "최종수정일",   align=S.ALIGN_CENTER,      fill=S.FILL_HEADER)  # F:G 병합
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        self._wc(ws, r, 8,  "",             align=S.ALIGN_CENTER)  # 최종수정일 값
        self._wc(ws, r, 9,  "엔티티분류",   align=S.ALIGN_CENTER,      fill=S.FILL_HEADER)
        self._wc(ws, r, 10, "",             align=S.ALIGN_CENTER)
        self._wc(ws, r, 11, "오너쉽",       align=S.ALIGN_CENTER,      fill=S.FILL_HEADER)
        self._wc(ws, r, 12, "",             align=S.ALIGN_CENTER)
        r += 1

        # Row 3: Entity 정의 | [B:L 병합 빈칸] (높이 120)
        ws.row_dimensions[r].height = S.ROW_HEIGHT_ENTITY_DEF
        self._wc(ws, r, 1, "Entity 정의", align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        self._wc(ws, r, 2, "",             align=S.ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
        r += 1

        return r

    # ──────────────────────────────────────────────────────────────────────────
    # 섹션 2: Key List (고정 8행)
    # ──────────────────────────────────────────────────────────────────────────

    def _write_key_list(self, ws: "Worksheet", row: int, table: TableDef) -> int:
        r = row

        # Row 4: 헤더 (높이 40)
        ws.row_dimensions[r].height = S.ROW_HEIGHT_KEY_HEADER
        self._wc(ws, r, 1, "구분\nKey List", align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        self._wc(ws, r, 2, "Key Name",        align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        self._wc(ws, r, 4, "Column Name",     align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=12)
        r += 1

        # Row 5: Primary Key — B:C = pk_constraint_name, D:L = pk_columns
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        pk_name = table.pk_constraint_name
        pk_cols = ", ".join(table.pk_columns) if table.pk_columns else ""
        self._wc(ws, r, 1, "Primary Key", align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        self._wc(ws, r, 2, pk_name,        align=S.ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        self._wc(ws, r, 4, pk_cols,        align=S.ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=12)
        r += 1

        # Row 6: Foreign Key — B:C = constraint_name (줄바꿈), D:L = "col → ref_table" (줄바꿈)
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        if table.fk_list:
            seen: set[str] = set()
            fk_names: list[str] = []
            fk_cols: list[str] = []
            for fk in table.fk_list:
                key = f"{fk.column}→{fk.ref_table}"
                if key not in seen:
                    seen.add(key)
                    fk_names.append(fk.constraint_name)
                    fk_cols.append(f"{fk.column} → {fk.ref_table}")
            fk_name_val = ", ".join(fk_names)
            fk_col_val = ", ".join(fk_cols)
        else:
            fk_name_val = ""
            fk_col_val = ""
        self._wc(ws, r, 1, "Foreign Key", align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        self._wc(ws, r, 2, fk_name_val,    align=S.ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        self._wc(ws, r, 4, fk_col_val,     align=S.ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=12)
        r += 1

        # Row 7: Index Key — B:C = idx.name (줄바꿈), D:L = ", ".join(idx.columns) (줄바꿈)
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        if table.indexes:
            idx_names = []
            idx_cols = []
            for idx in table.indexes:
                idx_names.append(idx.name)
                idx_cols.append(", ".join(idx.columns))
            idx_name_val = ", ".join(idx_names)
            idx_col_val = ", ".join(idx_cols)
        else:
            idx_name_val = ""
            idx_col_val = ""
        self._wc(ws, r, 1, "Index Key",  align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        self._wc(ws, r, 2, idx_name_val, align=S.ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        self._wc(ws, r, 4, idx_col_val,  align=S.ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=12)
        r += 1

        # Row 8: Sequence Key (빈칸)
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        self._wc(ws, r, 1, "Sequence Key", align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        self._wc(ws, r, 2, "",              align=S.ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        self._wc(ws, r, 4, "",              align=S.ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=12)
        r += 1

        # Row 9: Partitions Key 헤더 (빈칸)
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        self._wc(ws, r, 1, "Partitions Key",   align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        self._wc(ws, r, 4, "Partition Name",   align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        self._wc(ws, r, 9, "Partitions Value", align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=12)
        r += 1

        # Row 10: Partitions 값 행 (빈칸)
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        self._wc(ws, r, 1, "", align=S.ALIGN_CENTER)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        self._wc(ws, r, 4, "", align=S.ALIGN_CENTER)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        self._wc(ws, r, 9, "", align=S.ALIGN_CENTER)
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=12)
        r += 1

        # Row 11: SubPartitions Key (빈칸)
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        self._wc(ws, r, 1, "SubPartitions Key",   align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        self._wc(ws, r, 4, "",                     align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        self._wc(ws, r, 9, "SubPartitions Value",  align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=12)
        r += 1

        return r

    # ──────────────────────────────────────────────────────────────────────────
    # 섹션 3: 컬럼 헤더 + 데이터
    # ──────────────────────────────────────────────────────────────────────────

    def _write_columns(self, ws: "Worksheet", row: int, table: TableDef) -> int:
        r = row

        # Row 12: 컬럼 헤더 Row 1
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        single_headers = [
            (1, "No."),
            (2, "COLUMN NAME"),
            (3, "ATTRIBUTE NAME"),
            (4, "Type"),
            (5, "Length"),
            (6, "Null"),
            (7, "Keys"),
            (8, "인포타입명"),
            (9, "Description"),
            (12, "Source"),
            (13, "COMMENT SQL"),
        ]
        for col_idx, label in single_headers:
            self._wc(ws, r, col_idx, label, align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        # J:K 병합 → 'Attribute Info'
        self._wc(ws, r, 10, "Attribute Info", align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
        r += 1

        # Row 13: 컬럼 헤더 Row 2 (서브헤더)
        ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT
        self._wc(ws, r, 10, "Attribute Type",   align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        self._wc(ws, r, 11, "Relation & Value",  align=S.ALIGN_CENTER, fill=S.FILL_HEADER)
        # A~I, L, M: 위 행(r-1)과 병합
        for c in (1, 2, 3, 4, 5, 6, 7, 8, 9, 12, 13):
            ws.merge_cells(start_row=r - 1, start_column=c, end_row=r, end_column=c)
        r += 1

        # 데이터 행
        for col in table.columns:
            ws.row_dimensions[r].height = S.ROW_HEIGHT_DEFAULT

            # G열 Keys: PK / UK / FK 복합 표시
            keys_parts = []
            if col.is_pk:
                keys_parts.append("PK")
            if col.is_uk:
                keys_parts.append("UK")
            if col.fk_info:
                keys_parts.append("FK")
            keys_str = ",".join(keys_parts)

            # 기본 폰트: PK 행은 bold
            base_font = S.FONT_PK if col.is_pk else S.FONT_NORMAL

            null_str  = "NN" if col.not_null else ""
            attr_type = ""
            rel_val   = ""
            src_val   = ""
            if col.fk_info:
                attr_type = "RELATION"
                ref_cols  = ", ".join(col.fk_info.ref_columns) if col.fk_info.ref_columns else ""
                rel_val   = f"{col.fk_info.ref_table}.{ref_cols}" if ref_cols else col.fk_info.ref_table
            rule = _COLUMN_RULES.get((col.name, col.attribute_name))
            if rule:
                attr_type, rel_val = rule
            # dtl_code.csv: B열(col.name) 매칭
            if col.name in _DTL_BY_NAME:
                code_vals = _DTL_BY_NAME[col.name]
                attr_type = "코드 그룹"
                rel_val   = f'IND_CD="{col.name}"'
                src_val   = str(code_vals)
            # dtl_code.csv: C열(col.attribute_name) 매칭
            elif col.attribute_name and col.attribute_name in _DTL_BY_ATTR:
                code_group, code_vals = _DTL_BY_ATTR[col.attribute_name]
                attr_type = "코드 그룹"
                rel_val   = f'IND_CD="{code_group}"'
                src_val   = str(code_vals)

            # M열: AI 추론 attribute_name이 있는 경우 COMMENT SQL 생성
            comment_sql = ""
            if col.attribute_name_ai:
                comment_sql = (
                    f"COMMENT ON COLUMN {table.name}.{col.name} "
                    f"IS '{col.attribute_name}';"
                )

            # C열 / M열 폰트: AI 추론값이면 주황색
            attr_font = S.FONT_AI_SUGGESTED if col.attribute_name_ai else base_font

            row_data = [
                (1,  col.no,              S.ALIGN_CENTER,       base_font),
                (2,  col.name,            S.ALIGN_LEFT_NO_WRAP, base_font),
                (3,  col.attribute_name,  S.ALIGN_LEFT_NO_WRAP, attr_font),
                (4,  col.type_str,        S.ALIGN_CENTER,       base_font),
                (5,  col.length,          S.ALIGN_CENTER,       base_font),
                (6,  null_str,            S.ALIGN_CENTER,       base_font),
                (7,  keys_str,            S.ALIGN_CENTER,       base_font),
                (8,  "",                  S.ALIGN_LEFT_NO_WRAP, base_font),
                (9,  "",                  S.ALIGN_LEFT_NO_WRAP, base_font),
                (10, attr_type,           S.ALIGN_CENTER,       base_font),
                (11, rel_val,             S.ALIGN_LEFT_NO_WRAP, base_font),
                (12, src_val,             S.ALIGN_LEFT_NO_WRAP, base_font),
                (13, comment_sql,         S.ALIGN_LEFT_NO_WRAP, S.FONT_AI_SUGGESTED if comment_sql else base_font),
            ]
            for c_idx, val, align, font in row_data:
                self._wc(ws, r, c_idx, val, font=font, align=align)
            r += 1

        return r

    # ──────────────────────────────────────────────────────────────────────────
    # 컬럼 너비 적용
    # ──────────────────────────────────────────────────────────────────────────

    def _apply_column_widths(self, ws: "Worksheet") -> None:
        for col_letter, width in S.COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

    # ──────────────────────────────────────────────────────────────────────────
    # 헬퍼
    # ──────────────────────────────────────────────────────────────────────────

    def _wc(
        self,
        ws: "Worksheet",
        row: int,
        col: int,
        value,
        *,
        bold: bool = False,
        font: Font | None = None,
        align: Alignment | None = None,
        fill: PatternFill | None = None,
    ) -> None:
        """셀 값 + 폰트 + 정렬 + 테두리 + 배경색 적용."""
        cell = ws.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
        elif fill is S.FILL_HEADER:
            cell.font = S.FONT_HEADER_ON_DARK
        elif bold:
            cell.font = S.FONT_HEADER
        else:
            cell.font = S.FONT_NORMAL
        if align:
            cell.alignment = align
        cell.border = S.BORDER_THIN
        if fill:
            cell.fill = fill

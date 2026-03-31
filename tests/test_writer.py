import openpyxl
from pg_tabledef.writer.excel import ExcelWriter


class TestSheetStructure:
    def test_single_sheet_named_tabledef(self, sample_tables, tmp_path):
        """시트 1개, 이름 = '테이블정의서'."""
        out = tmp_path / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        wb = openpyxl.load_workbook(str(out))
        assert wb.sheetnames == ["테이블정의서"]

    def test_output_dir_auto_created(self, sample_tables, tmp_path):
        """output 폴더 없어도 자동 생성."""
        out = tmp_path / "new_dir" / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        assert out.exists()


class TestTableHeader:
    def test_table_name_in_sheet(self, sample_tables, tmp_path):
        """시트 어딘가에 테이블명이 존재."""
        out = tmp_path / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        ws = openpyxl.load_workbook(str(out))["테이블정의서"]
        all_values = [ws.cell(r, c).value for r in range(1, ws.max_row + 1) for c in range(1, 5)]
        assert "orders" in all_values

    def test_table_comment_in_sheet(self, sample_tables, tmp_path):
        out = tmp_path / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        ws = openpyxl.load_workbook(str(out))["테이블정의서"]
        all_values = [ws.cell(r, c).value for r in range(1, ws.max_row + 1) for c in range(1, 5)]
        assert "주문정보" in all_values


class TestColumnSection:
    def test_pk_marker_present(self, sample_tables, tmp_path):
        """PK 컬럼의 Keys 셀(G열)에 'PK' 값."""
        out = tmp_path / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        ws = openpyxl.load_workbook(str(out))["테이블정의서"]
        all_values = {ws.cell(r, c).value for r in range(1, ws.max_row + 1) for c in range(1, 13)}
        assert "PK" in all_values

    def test_fk_relation_value(self, sample_tables, tmp_path):
        """FK 컬럼의 Relation & Value(K열)에 'customers' 포함."""
        out = tmp_path / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        ws = openpyxl.load_workbook(str(out))["테이블정의서"]
        col_k_values = [ws.cell(r, 11).value for r in range(1, ws.max_row + 1)]
        assert any(v and "customers" in str(v) for v in col_k_values)

    def test_type_abbreviation_in_cell(self, sample_tables, tmp_path):
        """타입 셀(D열)에 약식 표기(VC, INT 등) 사용."""
        out = tmp_path / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        ws = openpyxl.load_workbook(str(out))["테이블정의서"]
        col_d_values = {ws.cell(r, 4).value for r in range(1, ws.max_row + 1)}
        assert "VC" in col_d_values or "INT" in col_d_values

    def test_not_null_shows_nn(self, sample_tables, tmp_path):
        """NOT NULL 컬럼의 Null 셀(F열)에 'NN' 표시."""
        out = tmp_path / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        ws = openpyxl.load_workbook(str(out))["테이블정의서"]
        col_f_values = {ws.cell(r, 6).value for r in range(1, ws.max_row + 1)}
        assert "NN" in col_f_values


class TestTableSeparation:
    def test_two_tables_both_present(self, sample_tables, tmp_path):
        """테이블 2개의 이름이 모두 시트에 존재."""
        out = tmp_path / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        ws = openpyxl.load_workbook(str(out))["테이블정의서"]
        all_values = {
            ws.cell(r, c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, 5)
        }
        assert "customers" in all_values
        assert "orders" in all_values


class TestTableCompleteness:
    def test_table_count_matches(self, sample_tables, tmp_path):
        """Excel에 출력된 테이블 수 == 파싱된 테이블 수."""
        out = tmp_path / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        ws = openpyxl.load_workbook(str(out))["테이블정의서"]
        excel_table_count = sum(
            1 for r in range(1, ws.max_row + 1)
            if ws.cell(r, 1).value == "Table 명"
        )
        assert excel_table_count == len(sample_tables)

    def test_column_count_per_table(self, sample_tables, tmp_path):
        """각 테이블의 Excel 출력 컬럼 행 수 == 파싱된 컬럼 수."""
        out = tmp_path / "out.xlsx"
        ExcelWriter(out).write(sample_tables)
        ws = openpyxl.load_workbook(str(out))["테이블정의서"]

        # "Table 명" 레이블 행 위치와 테이블명 수집
        table_header_rows = []
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "Table 명":
                table_header_rows.append((r, ws.cell(r, 2).value))
        table_header_rows.append((ws.max_row + 1, None))  # sentinel

        # 각 테이블 구간에서 A열이 정수인 행(컬럼 데이터 행) 카운트
        for i, (start_row, tname) in enumerate(table_header_rows[:-1]):
            end_row = table_header_rows[i + 1][0]
            excel_col_count = sum(
                1 for r in range(start_row, end_row)
                if isinstance(ws.cell(r, 1).value, int)
            )
            expected = next(t for t in sample_tables if t.name == tname)
            assert excel_col_count == len(expected.columns), (
                f"Table '{tname}': expected {len(expected.columns)} columns, "
                f"got {excel_col_count}"
            )
